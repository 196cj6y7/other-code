import os
from pathlib import Path
import easyocr
import pandas as pd
import openpyxl

def take_jpg_to_list(path):
    # Get the file name and number of pdf and save in pdf_list[0]
    directory = Path(path)
    files = directory.glob("*.jpeg")
    jpg_list = [file.name for file in files]
    print("JPG data list has been constructed.")
    return jpg_list
    

def read_jpg(data_list):   
    # Read the string in jpg file in English and turn them into string list(result)
    
    reader = easyocr.Reader(['en'],gpu=True)
    result = reader.readtext(data_list, detail = 0)
    return result
    
    
def create_dataframe(data_list, header_list, split_num):
    # df column size
    
    size = len(header_list)
    df = pd.DataFrame([data_list[i:i+size] for i in range(0, len(data_list), split_num)])  
    df.columns = header
    return df
    
def save_to_excel(df, sheet, file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    
    # check sheet name, if the sheet name is exist then rewrite the content of that sheet 
    # else create a new sheet with the sheet name
    ss = wb.sheetnames
    
    
    
    
    
    
    df.to_excel(file_name, sheet_name=sheet, index=True)
    wb.save(file_name)
    return
            
           
#======================== main ====================================================

# create a DataFrame to collect the data from JPG list
header = ["Item No.", "Purchase Order No.", "Rev. No.", "KKS No.", "Category & Description", "Valve Type"]
data_to_excel = pd.DataFrame(columns=header)
column_number = 6

workbook = openpyxl.Workbook()
workbook.save('valve.xlsx')

#------------------------------------------------

jpg_file_name = take_jpg_to_list("")
for i in range(len(jpg_file_name)):
    jpg_content_list = read_jpg(jpg_file_name[i])
    jpg_dataframe = create_dataframe(jpg_content_list, header, column_number)
    save_to_excel(jpg_dataframe, jpg_file_name[i], 'valve.xlsx')
    
    


